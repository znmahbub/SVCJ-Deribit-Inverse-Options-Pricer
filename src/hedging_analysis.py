from __future__ import annotations

from dataclasses import dataclass, field
from functools import lru_cache
from pathlib import Path
from typing import Any
import math
import re

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from src.calibration import _build_pricing_plan, _price_with_plan
from src.inverse_fft_pricer import FFTParams, cf_heston, cf_svcj

try:
    from tqdm.auto import tqdm
except Exception:  # pragma: no cover
    tqdm = None


MODELS = ("black", "heston", "svcj")
HEDGE_TYPES = ("delta", "net_delta")
PARAM_SHEETS = {
    "black": "black_params",
    "heston": "heston_params",
    "svcj": "svcj_params",
}
PRICE_COLS = {
    "black": "price_black",
    "heston": "price_heston",
    "svcj": "price_svcj",
}
DELTA_OUTPUT_COLS = [
    "regular_delta_black",
    "net_delta_black",
    "regular_delta_heston",
    "net_delta_heston",
    "regular_delta_svcj",
    "net_delta_svcj",
]
TS_RE = re.compile(r"deribit_options_snapshot_(\d{8}T\d{6})Z\.csv$")
EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_DATA_ROWS = EXCEL_MAX_ROWS - 1


@dataclass(frozen=True)
class HedgingConfig:
    calibration_xlsx: Path
    data_dir: Path
    output_xlsx: Path
    hedge_price_file: str | None = None
    funding_file: str | None = None
    dynamic_b: bool = True
    verify_prices: bool = True
    delta_u_max: float = 200.0
    delta_quad_n: int = 128
    rebalance_every_n_snapshots: int = 1
    max_snapshot_groups: int | None = None
    currencies: tuple[str, ...] | None = None
    exclude_expiry_crossing: bool = True
    fft_params: FFTParams = field(
        default_factory=lambda: FFTParams(
            N=2**12,
            alpha=1.5,
            eta=0.10,
            b=-10.0,
            use_simpson=True,
        )
    )

    def __post_init__(self) -> None:
        if int(self.rebalance_every_n_snapshots) < 1:
            raise ValueError("rebalance_every_n_snapshots must be >= 1")


def _to_utc_ts(x: Any) -> pd.Timestamp:
    return pd.to_datetime(x, utc=True)


def _iso_z(x: Any) -> str:
    return _to_utc_ts(x).strftime("%Y-%m-%dT%H:%M:%SZ")


def _num(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce")


def _resolve_data_file(data_dir: Path, preferred: str | None, candidates: tuple[str, ...]) -> tuple[Path, str]:
    if preferred is not None:
        path = data_dir / preferred
        if not path.exists():
            raise FileNotFoundError(f"Missing required file: {path}")
        return path, preferred
    for name in candidates:
        path = data_dir / name
        if path.exists():
            return path, name
    raise FileNotFoundError(f"Could not find any of {candidates} under {data_dir}")


def _stringify_datetimes(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in out.columns:
        if pd.api.types.is_datetime64_any_dtype(out[col]):
            out[col] = out[col].map(lambda x: _iso_z(x) if pd.notna(x) else None)
    return out


def _make_progress(total: int, desc: str):
    if tqdm is not None:
        return tqdm(total=total, desc=desc, leave=False)

    class _FallbackProgress:
        def __init__(self, total: int, desc: str):
            self.total = max(int(total), 1)
            self.desc = desc
            self.n = 0

        def update(self, n: int = 1):
            self.n += n

        def set_description(self, desc: str):
            self.desc = desc

        def close(self):
            return None

    return _FallbackProgress(total, desc)


@lru_cache(maxsize=None)
def _leggauss_cached(n: int) -> tuple[np.ndarray, np.ndarray]:
    return np.polynomial.legendre.leggauss(int(n))


def _norm_cdf(x: np.ndarray | float) -> np.ndarray | float:
    x_arr = np.asarray(x, dtype=float)
    cdf = 0.5 * (1.0 + np.vectorize(math.erf)(x_arr / math.sqrt(2.0)))
    return float(cdf) if np.isscalar(x) else cdf


def load_calibration_workbook(path: Path) -> dict[str, pd.DataFrame]:
    needed = [*PARAM_SHEETS.values(), "train_data", "test_data"]
    return {sheet: pd.read_excel(path, sheet_name=sheet, engine="openpyxl") for sheet in needed}


def combine_option_sheets(workbook: dict[str, pd.DataFrame]) -> pd.DataFrame:
    train = workbook["train_data"].copy()
    test = workbook["test_data"].copy()
    train["split"] = "train"
    test["split"] = "test"
    df = pd.concat([train, test], ignore_index=True, sort=False)

    df["snapshot_ts"] = pd.to_datetime(df["snapshot_ts"], utc=True)
    if "expiry_datetime" in df.columns:
        df["expiry_datetime"] = pd.to_datetime(df["expiry_datetime"], utc=True, errors="coerce")
    if "timestamp" in df.columns:
        df["timestamp"] = pd.to_datetime(df["timestamp"], utc=True, errors="coerce")

    df["currency"] = df["currency"].astype(str).str.upper()
    df["option_type"] = df["option_type"].astype(str).str.lower()

    numeric_cols = [
        "strike", "time_to_maturity", "mid_price_clean", "price_black", "price_heston", "price_svcj",
        "F0", "futures_price", "moneyness", "log_moneyness", "delta",
    ]
    for c in numeric_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    if "F0" not in df.columns and "futures_price" in df.columns:
        df["F0"] = _num(df["futures_price"])
    if "mid_price_clean" not in df.columns:
        if {"bid_price", "ask_price"}.issubset(df.columns):
            bid = _num(df["bid_price"])
            ask = _num(df["ask_price"])
            valid = np.isfinite(bid) & np.isfinite(ask) & (ask >= bid)
            df["mid_price_clean"] = np.where(valid, 0.5 * (bid + ask), np.nan)
        elif "mid_price" in df.columns:
            df["mid_price_clean"] = _num(df["mid_price"])
        else:
            raise ValueError("Could not infer current option mid price.")

    return df.sort_values(["snapshot_ts", "currency", "instrument_name"]).reset_index(drop=True)


def filter_options_for_run(options: pd.DataFrame, cfg: HedgingConfig) -> pd.DataFrame:
    out = options.copy()
    if cfg.currencies:
        allowed = {str(x).upper() for x in cfg.currencies}
        out = out[out["currency"].isin(allowed)].copy()
    if cfg.max_snapshot_groups is not None:
        unique_ts = np.sort(out["snapshot_ts"].dropna().unique())
        keep_ts = set(unique_ts[: int(cfg.max_snapshot_groups)])
        out = out[out["snapshot_ts"].isin(keep_ts)].copy()
    return out.sort_values(["snapshot_ts", "currency", "instrument_name"]).reset_index(drop=True)


def build_param_lookup(workbook: dict[str, pd.DataFrame]) -> dict[tuple[str, str, str], dict[str, Any]]:
    lookup: dict[tuple[str, str, str], dict[str, Any]] = {}
    for model, sheet in PARAM_SHEETS.items():
        df = workbook[sheet].copy()
        df["timestamp"] = pd.to_datetime(df["timestamp"], utc=True)
        df["currency"] = df["currency"].astype(str).str.upper()
        for _, row in df.iterrows():
            lookup[(model, _iso_z(row["timestamp"]), row["currency"])] = row.to_dict()
    return lookup


def list_snapshot_files(data_dir: Path) -> pd.DataFrame:
    rows = []
    for path in sorted(data_dir.glob("deribit_options_snapshot_*.csv")):
        match = TS_RE.search(path.name)
        if match:
            rows.append({
                "snapshot_ts": pd.to_datetime(match.group(1), format="%Y%m%dT%H%M%S", utc=True),
                "csv_path": path,
            })
    out = pd.DataFrame(rows).sort_values("snapshot_ts").reset_index(drop=True)
    if out.empty:
        raise FileNotFoundError(f"No deribit_options_snapshot_*.csv found under {data_dir}")
    return out


def build_forward_snapshot_lookup(snapshot_index: pd.DataFrame, step_n: int = 1) -> dict[str, dict[str, Any]]:
    ts_list = snapshot_index["snapshot_ts"].tolist()
    path_list = snapshot_index["csv_path"].tolist()
    lookup: dict[str, dict[str, Any]] = {}
    for i, ts in enumerate(ts_list):
        j = i + int(step_n)
        eval_ts = ts_list[j] if j < len(ts_list) else pd.NaT
        eval_path = path_list[j] if j < len(path_list) else None
        lookup[_iso_z(ts)] = {
            "eval_snapshot_ts": eval_ts,
            "eval_csv_path": eval_path,
        }
    return lookup


@lru_cache(maxsize=512)
def _load_snapshot_minimal(snapshot_path: str) -> pd.DataFrame:
    df = pd.read_csv(snapshot_path)
    if "currency" in df.columns:
        df["currency"] = df["currency"].astype(str).str.upper()
    elif "coin" in df.columns:
        df["currency"] = df["coin"].astype(str).str.upper()
    else:
        raise ValueError(f"Snapshot file {snapshot_path} missing currency/coin column")
    if "instrument_name" not in df.columns:
        raise ValueError(f"Snapshot file {snapshot_path} missing instrument_name column")

    if "market_price" in df.columns:
        market_price = _num(df["market_price"])
    elif {"bid_price", "ask_price"}.issubset(df.columns):
        bid = _num(df["bid_price"])
        ask = _num(df["ask_price"])
        valid = np.isfinite(bid) & np.isfinite(ask) & (ask >= bid)
        market_price = pd.Series(np.where(valid, 0.5 * (bid + ask), np.nan), index=df.index)
    elif "mid_price" in df.columns:
        market_price = _num(df["mid_price"])
    else:
        raise ValueError(f"Snapshot file {snapshot_path} missing market_price-compatible columns")

    if "futures_price" in df.columns:
        F0 = _num(df["futures_price"])
    elif "forward_price" in df.columns:
        F0 = _num(df["forward_price"])
    elif "F0" in df.columns:
        F0 = _num(df["F0"])
    else:
        F0 = pd.Series(np.nan, index=df.index)

    out = df[["currency", "instrument_name"]].copy()
    out["market_price_next"] = market_price
    out["F0_next"] = F0
    out["time_to_maturity_next"] = _num(df["time_to_maturity"]) if "time_to_maturity" in df.columns else np.nan
    return out


def load_perp_history(data_dir: Path, filename: str | None = None) -> tuple[pd.DataFrame, str]:
    path, used = _resolve_data_file(
        data_dir,
        filename,
        ("perpetual_futures_prices_history.csv", "perpetual_futures_prices.csv"),
    )
    df = pd.read_csv(path)

    # New updater format:
    #   obs_datetime, coin, perp_futures_mark_price, perp_futures_best_bid, perp_futures_best_ask, ...
    if {"obs_datetime", "coin", "perp_futures_mark_price"}.issubset(df.columns):
        out = df.copy()
        out["currency"] = out["coin"].astype(str).str.upper()
        out["timestamp"] = pd.to_datetime(out["obs_datetime"], utc=True, errors="coerce")
        out["close_price"] = pd.to_numeric(out["perp_futures_mark_price"], errors="coerce")

        # Fallback only if mark price is missing.
        if "perp_futures_best_bid" in out.columns and "perp_futures_best_ask" in out.columns:
            bid = pd.to_numeric(out["perp_futures_best_bid"], errors="coerce")
            ask = pd.to_numeric(out["perp_futures_best_ask"], errors="coerce")
            valid_mid = np.isfinite(bid) & np.isfinite(ask) & (ask >= bid)
            missing_close = ~np.isfinite(out["close_price"])
            out.loc[missing_close & valid_mid, "close_price"] = 0.5 * (bid[missing_close & valid_mid] + ask[missing_close & valid_mid])

        if "trade_price" in out.columns:
            trade_price = pd.to_numeric(out["trade_price"], errors="coerce")
            missing_close = ~np.isfinite(out["close_price"])
            out.loc[missing_close, "close_price"] = trade_price[missing_close]

        out = out.dropna(subset=["currency", "timestamp", "close_price"]).copy()
        out = out.sort_values(["currency", "timestamp"]).drop_duplicates(subset=["currency", "timestamp"], keep="first").reset_index(drop=True)
        if out.empty:
            raise ValueError(f"{path} contains the new perpetual futures format, but no usable rows after parsing.")
        return out[["currency", "timestamp", "close_price"]], used

    raise ValueError(
        f"{path} must contain the new perpetual futures columns "
        "(obs_datetime, coin, perp_futures_mark_price)."
    )


def load_funding_history(data_dir: Path, filename: str | None = None) -> tuple[pd.DataFrame, str | None]:
    candidates = ("perpetual_funding_history.csv", "perpetual_funding.csv")
    chosen = None
    if filename is not None:
        chosen = filename
    else:
        for name in candidates:
            if (data_dir / name).exists():
                chosen = name
                break
    if chosen is None:
        return pd.DataFrame(columns=["currency", "timestamp", "funding_rate"]), None
    path = data_dir / chosen
    df = pd.read_csv(path)
    if {"currency", "timestamp", "funding_rate"} - set(df.columns):
        return pd.DataFrame(columns=["currency", "timestamp", "funding_rate"]), chosen
    df["currency"] = df["currency"].astype(str).str.upper()
    df["timestamp"] = pd.to_datetime(df["timestamp"], utc=True, errors="coerce")
    df["funding_rate"] = pd.to_numeric(df["funding_rate"], errors="coerce")
    df = df.dropna(subset=["currency", "timestamp"])
    return df.sort_values(["currency", "timestamp"]).reset_index(drop=True), chosen


def _black_call_delta(F: np.ndarray, K: np.ndarray, T: np.ndarray, sigma: np.ndarray) -> np.ndarray:
    out = np.full_like(F, np.nan, dtype=float)
    valid = np.isfinite(F) & np.isfinite(K) & np.isfinite(T) & np.isfinite(sigma) & (F > 0) & (K > 0) & (T > 0) & (sigma > 0)
    if not np.any(valid):
        return out
    srt = sigma[valid] * np.sqrt(T[valid])
    d1 = (np.log(F[valid] / K[valid]) + 0.5 * sigma[valid] ** 2 * T[valid]) / srt
    out[valid] = _norm_cdf(d1)
    return out


def _intrinsic_coin_price(option_type: np.ndarray, F: np.ndarray, K: np.ndarray) -> np.ndarray:
    is_call = np.char.lower(option_type.astype(str)) == "call"
    return np.where(is_call, np.maximum(1.0 - K / F, 0.0), np.maximum(K / F - 1.0, 0.0))


def _intrinsic_regular_delta(option_type: np.ndarray, F: np.ndarray, K: np.ndarray) -> np.ndarray:
    is_call = np.char.lower(option_type.astype(str)) == "call"
    return np.where(is_call, (F > K).astype(float), -(K > F).astype(float))


def _model_prices_from_plan(model: str, block: pd.DataFrame, params: dict[str, Any], fft_params: FFTParams, dynamic_b: bool = True, use_cache: bool = True) -> np.ndarray:
    if block.empty:
        return np.array([], dtype=float)
    out = np.full(len(block), np.nan, dtype=float)
    F = _num(block["F0"]).to_numpy(dtype=float)
    K = _num(block["strike"]).to_numpy(dtype=float)
    T = _num(block["time_to_maturity"]).to_numpy(dtype=float)
    opt_type = block["option_type"].astype(str).str.lower().to_numpy()
    valid = np.isfinite(F) & (F > 0) & np.isfinite(K) & (K > 0) & np.isfinite(T) & (T >= 0)
    expired = valid & (T <= 0)
    if np.any(expired):
        out[expired] = _intrinsic_coin_price(opt_type[expired], F[expired], K[expired])
    live = valid & (T > 0)
    if not np.any(live):
        return out
    live_block = block.loc[live].copy()
    live_prices = np.full(len(live_block), np.nan, dtype=float)
    pricing_plan = _build_pricing_plan(live_block, fft_params_base=fft_params, dynamic_b=dynamic_b, fft_params_by_expiry=None)
    _price_with_plan(pricing_plan, model=model.lower(), params=params, out_prices=live_prices, use_cache=use_cache)
    out[np.flatnonzero(live)] = live_prices
    return out


def _semi_analytic_call_p1_from_logprice_cf(K: np.ndarray, cf, quad_u_max: float = 200.0, quad_n: int = 128) -> np.ndarray:
    K = np.asarray(K, dtype=float)
    out = np.full_like(K, np.nan)
    valid = np.isfinite(K) & (K > 0)
    if not np.any(valid):
        return out
    K_valid = K[valid]
    lnK = np.log(K_valid)
    nodes, weights = _leggauss_cached(int(quad_n))
    u = 0.5 * quad_u_max * (nodes + 1.0)
    w = 0.5 * quad_u_max * weights
    phi_u_minus_i = np.asarray(cf(u - 1j), dtype=np.complex128).reshape(-1)
    phi_minus_i = np.asarray(cf(np.array([-1j])), dtype=np.complex128).reshape(-1)[0]
    exp_term = np.exp(-1j * u[:, None] * lnK[None, :])
    integrand = exp_term * (phi_u_minus_i[:, None] / ((1j * u)[:, None] * phi_minus_i))
    p1 = 0.5 + (1.0 / np.pi) * np.sum(w[:, None] * np.real(integrand), axis=0)
    out[valid] = np.clip(p1, 0.0, 1.0)
    return out


def _regular_delta_block(model: str, block: pd.DataFrame, params: dict[str, Any], fft_params: FFTParams, dynamic_b: bool = True, delta_u_max: float = 200.0, delta_quad_n: int = 128) -> np.ndarray:
    if block.empty:
        return np.array([], dtype=float)
    out = np.full(len(block), np.nan, dtype=float)
    F = _num(block["F0"]).to_numpy(dtype=float)
    K = _num(block["strike"]).to_numpy(dtype=float)
    T = _num(block["time_to_maturity"]).to_numpy(dtype=float)
    opt_type = block["option_type"].astype(str).str.lower().to_numpy()
    valid = np.isfinite(F) & (F > 0) & np.isfinite(K) & (K > 0) & np.isfinite(T) & (T >= 0)
    expired = valid & (T <= 0)
    if np.any(expired):
        out[expired] = _intrinsic_regular_delta(opt_type[expired], F[expired], K[expired])
    live = valid & (T > 0)
    if not np.any(live):
        return out
    live_block = block.loc[live].copy()
    pricing_plan = _build_pricing_plan(live_block, fft_params_base=fft_params, dynamic_b=dynamic_b, fft_params_by_expiry=None)
    live_positions = np.flatnonzero(live)
    for idx, K_g, put_mask, T_g, F0_g, _ in pricing_plan:
        K_g = np.asarray(K_g, dtype=float)
        if model == "black":
            p1 = _black_call_delta(np.full_like(K_g, float(F0_g)), K_g, np.full_like(K_g, float(T_g)), np.full_like(K_g, float(params["sigma"])))
        elif model == "heston":
            cf = lambda u: cf_heston(u, T=float(T_g), F0=float(F0_g), kappa=float(params["kappa"]), theta=float(params["theta"]), sigma_v=float(params["sigma_v"]), rho=float(params["rho"]), v0=float(params["v0"]))
            p1 = _semi_analytic_call_p1_from_logprice_cf(K_g, cf=cf, quad_u_max=delta_u_max, quad_n=delta_quad_n)
        elif model == "svcj":
            cf = lambda u: cf_svcj(u, T=float(T_g), F0=float(F0_g), kappa=float(params["kappa"]), theta=float(params["theta"]), sigma_v=float(params["sigma_v"]), rho=float(params["rho"]), v0=float(params["v0"]), lam=float(params["lam"]), ell_y=float(params["ell_y"]), sigma_y=float(params["sigma_y"]), ell_v=float(params["ell_v"]), rho_j=float(params["rho_j"]))
            p1 = _semi_analytic_call_p1_from_logprice_cf(K_g, cf=cf, quad_u_max=delta_u_max, quad_n=delta_quad_n)
        else:
            raise ValueError(f"Unsupported model: {model}")
        reg = p1.copy()
        reg[put_mask] = reg[put_mask] - 1.0
        out[live_positions[idx]] = reg
    return out


def compute_prices_and_deltas(model: str, block: pd.DataFrame, params: dict[str, Any], fft_params: FFTParams, dynamic_b: bool = True, delta_u_max: float = 200.0, delta_quad_n: int = 128) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    prices = _model_prices_from_plan(model=model, block=block, params=params, fft_params=fft_params, dynamic_b=dynamic_b, use_cache=True)
    reg_delta = _regular_delta_block(model=model, block=block, params=params, fft_params=fft_params, dynamic_b=dynamic_b, delta_u_max=delta_u_max, delta_quad_n=delta_quad_n)
    return prices, reg_delta, reg_delta - prices


def enrich_options_with_model_outputs(options: pd.DataFrame, param_lookup: dict[tuple[str, str, str], dict[str, Any]], fft_params: FFTParams, dynamic_b: bool = True, verify_prices: bool = True, delta_u_max: float = 200.0, delta_quad_n: int = 128, progress=None) -> tuple[pd.DataFrame, pd.DataFrame]:
    out = options.copy()
    for model in MODELS:
        out[f"model_price_{model}"] = np.nan
        out[f"regular_delta_{model}"] = np.nan
        out[f"net_delta_{model}"] = np.nan

    rows: list[dict[str, Any]] = []
    for (snapshot_ts, currency), idx in out.groupby(["snapshot_ts", "currency"], sort=True).groups.items():
        block = out.loc[list(idx)].copy()
        ts_iso = _iso_z(snapshot_ts)
        for model in MODELS:
            params = param_lookup.get((model, ts_iso, currency))
            if params is None:
                rows.append({"snapshot_ts": ts_iso, "currency": currency, "model": model, "n_rows": int(len(block)), "n_valid_prices": 0, "n_valid_diff": 0, "mean_abs_price_diff": np.nan, "max_abs_price_diff": np.nan, "rmse_price_diff": np.nan, "all_nan_diff": True, "error": "missing_params"})
                if progress is not None:
                    progress.update(1)
                continue
            error_msg = None
            try:
                prices, reg_delta, net_delta = compute_prices_and_deltas(model=model, block=block, params=params, fft_params=fft_params, dynamic_b=dynamic_b, delta_u_max=delta_u_max, delta_quad_n=delta_quad_n)
                out.loc[list(idx), f"model_price_{model}"] = prices
                out.loc[list(idx), f"regular_delta_{model}"] = reg_delta
                out.loc[list(idx), f"net_delta_{model}"] = net_delta
            except Exception as exc:  # pragma: no cover
                prices = np.full(len(block), np.nan)
                reg_delta = np.full(len(block), np.nan)
                net_delta = np.full(len(block), np.nan)
                error_msg = str(exc)
            target = _num(block[PRICE_COLS[model]]).to_numpy(dtype=float) if verify_prices and PRICE_COLS[model] in block.columns else np.full(len(block), np.nan)
            diff = prices - target
            valid_diff = np.isfinite(diff)
            rows.append({
                "snapshot_ts": ts_iso,
                "currency": currency,
                "model": model,
                "n_rows": int(len(block)),
                "n_valid_prices": int(np.isfinite(prices).sum()),
                "n_valid_reg_delta": int(np.isfinite(reg_delta).sum()),
                "n_valid_net_delta": int(np.isfinite(net_delta).sum()),
                "n_valid_diff": int(valid_diff.sum()),
                "mean_abs_price_diff": float(np.mean(np.abs(diff[valid_diff]))) if valid_diff.any() else np.nan,
                "max_abs_price_diff": float(np.max(np.abs(diff[valid_diff]))) if valid_diff.any() else np.nan,
                "rmse_price_diff": float(np.sqrt(np.mean(diff[valid_diff] ** 2))) if valid_diff.any() else np.nan,
                "all_nan_diff": bool(not valid_diff.any()),
                "error": error_msg,
            })
            if progress is not None:
                progress.update(1)
    return out, pd.DataFrame(rows)


def attach_evaluation_snapshot_data(options: pd.DataFrame, snapshot_index: pd.DataFrame, rebalance_every_n_snapshots: int = 1) -> pd.DataFrame:
    out = options.copy()
    lookup = build_forward_snapshot_lookup(snapshot_index, step_n=rebalance_every_n_snapshots)
    out["snapshot_ts_iso"] = out["snapshot_ts"].map(_iso_z)
    out["eval_snapshot_ts"] = out["snapshot_ts_iso"].map(lambda x: lookup.get(x, {}).get("eval_snapshot_ts"))
    out["eval_snapshot_ts"] = pd.to_datetime(out["eval_snapshot_ts"], utc=True, errors="coerce")

    eval_market_price = np.full(len(out), np.nan)
    eval_F0 = np.full(len(out), np.nan)
    eval_ttm = np.full(len(out), np.nan)

    for ts_iso, row_ids in out.groupby("snapshot_ts_iso").groups.items():
        meta = lookup.get(ts_iso, {})
        eval_path = meta.get("eval_csv_path")
        if eval_path is None:
            continue
        nxt_df = _load_snapshot_minimal(str(eval_path))
        cur = out.loc[list(row_ids), ["currency", "instrument_name"]].copy()
        cur["__rid"] = list(row_ids)
        merged = cur.merge(nxt_df, on=["currency", "instrument_name"], how="left")
        rid = merged["__rid"].to_numpy(dtype=int)
        eval_market_price[rid] = _num(merged["market_price_next"]).to_numpy(dtype=float)
        eval_F0[rid] = _num(merged["F0_next"]).to_numpy(dtype=float)
        eval_ttm[rid] = _num(merged["time_to_maturity_next"]).to_numpy(dtype=float)

    out["eval_market_price"] = eval_market_price
    out["eval_F0"] = eval_F0
    out["eval_time_to_maturity"] = eval_ttm
    out["has_eval_market_price"] = np.isfinite(out["eval_market_price"])
    out["dt_years"] = (out["eval_snapshot_ts"] - out["snapshot_ts"]).dt.total_seconds() / (365.25 * 24.0 * 3600.0)
    out["expires_before_eval_snapshot"] = out["expiry_datetime"].notna() & out["eval_snapshot_ts"].notna() & (out["expiry_datetime"] <= out["eval_snapshot_ts"])

    # backward-compatible names
    out["next_snapshot_ts"] = out["eval_snapshot_ts"]
    out["next_market_price"] = out["eval_market_price"]
    out["next_F0"] = out["eval_F0"]
    out["next_time_to_maturity"] = out["eval_time_to_maturity"]
    out["has_next_market_price"] = out["has_eval_market_price"]
    out["expires_before_next_snapshot"] = out["expires_before_eval_snapshot"]
    return out


def attach_perp_and_funding(options: pd.DataFrame, perp: pd.DataFrame, funding: pd.DataFrame) -> pd.DataFrame:
    out = options.copy()
    t_prices = perp.rename(columns={"timestamp": "snapshot_ts", "close_price": "perp_close_t"})[["currency", "snapshot_ts", "perp_close_t"]]
    t1_prices = perp.rename(columns={"timestamp": "eval_snapshot_ts", "close_price": "perp_close_t1"})[["currency", "eval_snapshot_ts", "perp_close_t1"]]
    out = out.merge(t_prices, on=["currency", "snapshot_ts"], how="left")
    out = out.merge(t1_prices, on=["currency", "eval_snapshot_ts"], how="left")
    out["is_hedgeable_interval"] = (
        out["has_eval_market_price"]
        & np.isfinite(out["perp_close_t"])
        & np.isfinite(out["perp_close_t1"])
        & (out["perp_close_t"] > 0)
        & (out["perp_close_t1"] > 0)
        & np.isfinite(out["F0"])
        & (_num(out["F0"]) > 0)
    )
    out["cum_funding_rate"] = 0.0
    out["avg_funding_rate"] = 0.0
    out["n_funding_obs"] = 0
    if funding.empty:
        return out
    for currency, idx in out.groupby("currency").groups.items():
        row_idx = np.asarray(list(idx), dtype=int)
        sub = out.loc[row_idx, ["snapshot_ts", "eval_snapshot_ts"]]
        fund = funding[funding["currency"] == currency].copy()
        if fund.empty:
            continue
        ts = fund["timestamp"].to_numpy(dtype="datetime64[ns]")
        vals = _num(fund["funding_rate"]).fillna(0.0).to_numpy(dtype=float)
        cs = np.concatenate([[0.0], np.cumsum(vals)])
        cc = np.concatenate([[0], np.cumsum(np.ones_like(vals, dtype=int))])
        t0 = pd.to_datetime(sub["snapshot_ts"], utc=True, errors="coerce").to_numpy(dtype="datetime64[ns]")
        t1 = pd.to_datetime(sub["eval_snapshot_ts"], utc=True, errors="coerce").to_numpy(dtype="datetime64[ns]")
        valid = ~pd.isna(t0) & ~pd.isna(t1)
        left = np.searchsorted(ts, t0[valid], side="right")
        right = np.searchsorted(ts, t1[valid], side="right")
        sums = np.zeros(len(sub), dtype=float)
        counts = np.zeros(len(sub), dtype=int)
        sums[valid] = cs[right] - cs[left]
        counts[valid] = cc[right] - cc[left]
        means = np.divide(sums, counts, out=np.zeros_like(sums), where=counts > 0)
        out.loc[row_idx, "cum_funding_rate"] = sums
        out.loc[row_idx, "avg_funding_rate"] = means
        out.loc[row_idx, "n_funding_obs"] = counts
    return out


def build_hedge_interval_panel(options: pd.DataFrame, *, exclude_expiry_crossing: bool = True, rebalance_every_n_snapshots: int = 1) -> pd.DataFrame:
    base_cols = [
        "snapshot_ts", "eval_snapshot_ts", "next_snapshot_ts", "currency", "split", "instrument_name", "option_type", "strike",
        "expiry_datetime", "time_to_maturity", "eval_time_to_maturity", "next_time_to_maturity", "F0", "eval_F0", "next_F0",
        "moneyness", "log_moneyness", "perp_close_t", "perp_close_t1", "cum_funding_rate", "avg_funding_rate", "n_funding_obs",
        "is_hedgeable_interval", "expires_before_eval_snapshot", "expires_before_next_snapshot", "dt_years",
    ]
    panel = options[[c for c in base_cols if c in options.columns]].copy()
    opt_t = _num(options["mid_price_clean"])
    opt_t1 = _num(options["eval_market_price"])
    F_t = _num(options["F0"])
    H_t = _num(options["perp_close_t"])
    H_t1 = _num(options["perp_close_t1"])
    inverse_move = np.where(np.isfinite(H_t) & np.isfinite(H_t1) & (H_t > 0) & (H_t1 > 0), (1.0 / H_t) - (1.0 / H_t1), np.nan)
    panel["rebalance_every_n_snapshots"] = int(rebalance_every_n_snapshots)
    panel["option_price_coin_t"] = opt_t
    panel["option_price_coin_t1"] = opt_t1
    panel["option_pnl_coin_short"] = -(opt_t1 - opt_t)
    panel["is_summary_eligible"] = options["is_hedgeable_interval"].fillna(False)
    if exclude_expiry_crossing:
        panel["is_summary_eligible"] &= ~options["expires_before_eval_snapshot"].fillna(False)

    for model in MODELS:
        reg_delta = _num(options[f"regular_delta_{model}"])
        net_delta = _num(options[f"net_delta_{model}"])
        panel[f"hedge_notional_usd_{model}_delta"] = H_t * reg_delta
        panel[f"net_pnl_coin_{model}_delta"] = panel["option_pnl_coin_short"] + panel[f"hedge_notional_usd_{model}_delta"] * inverse_move
        net_scale = np.where(np.isfinite(F_t) & (F_t > 0), (H_t ** 2) / F_t, np.nan)
        panel[f"hedge_notional_usd_{model}_net_delta"] = net_scale * net_delta
        panel[f"net_pnl_coin_{model}_net_delta"] = panel["option_pnl_coin_short"] + panel[f"hedge_notional_usd_{model}_net_delta"] * inverse_move

    invalid = ~panel["is_summary_eligible"].fillna(False)
    metric_cols = [c for c in panel.columns if c.startswith("hedge_notional_usd_") or c.startswith("net_pnl_coin_")]
    panel.loc[invalid, metric_cols + ["option_pnl_coin_short"]] = np.nan
    return panel.sort_values(["snapshot_ts", "currency", "instrument_name"]).reset_index(drop=True)


def add_analysis_buckets(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    ttm_days = pd.to_numeric(out["time_to_maturity"], errors="coerce") * 365.25
    out["maturity_bucket"] = pd.cut(
        ttm_days,
        bins=[-np.inf, 7, 14, 30, 90, np.inf],
        labels=["0-7d", "7-14d", "14-30d", "30-90d", "90d+"],
    ).astype("object")
    if "moneyness" in out.columns and out["moneyness"].notna().any():
        m = pd.to_numeric(out["moneyness"], errors="coerce")
    elif "F0" in out.columns and "strike" in out.columns:
        m = pd.to_numeric(out["strike"], errors="coerce") / pd.to_numeric(out["F0"], errors="coerce")
    else:
        m = pd.Series(np.nan, index=out.index)
    out["moneyness_ratio"] = m
    out["moneyness_bucket"] = pd.cut(
        m,
        bins=[-np.inf, 0.95, 1.05, np.inf],
        labels=["OTM (K/F<0.95)", "ATM (0.95-1.05)", "ITM (K/F>1.05)"],
    ).astype("object")
    return out


def _tail_stats(x: pd.Series) -> tuple[float, float]:
    vals = pd.to_numeric(x, errors="coerce").dropna().to_numpy(dtype=float)
    if len(vals) == 0:
        return np.nan, np.nan
    q05 = float(np.quantile(vals, 0.05))
    tail = vals[vals <= q05]
    es05 = float(np.mean(tail)) if len(tail) else np.nan
    return q05, es05


def _summarize_errors(unhedged: pd.Series, hedged: pd.Series, hedge_notional: pd.Series) -> dict[str, Any]:
    paired = pd.DataFrame({"un": _num(unhedged), "he": _num(hedged), "notional": _num(hedge_notional)}).dropna(subset=["un", "he"])
    if paired.empty:
        return {
            "n_intervals": 0, "mean_unhedged_coin": np.nan, "mean_hedged_coin": np.nan, "std_unhedged_coin": np.nan,
            "std_hedged_coin": np.nan, "var_unhedged_coin": np.nan, "var_hedged_coin": np.nan,
            "rmse_unhedged_coin": np.nan, "rmse_hedged_coin": np.nan, "rmse_reduction": np.nan,
            "mae_unhedged_coin": np.nan, "mae_hedged_coin": np.nan, "mae_reduction": np.nan,
            "variance_reduction": np.nan, "hit_rate_abs_improvement": np.nan, "q05_hedged_coin": np.nan,
            "es05_hedged_coin": np.nan, "mean_abs_hedge_notional_usd": np.nan,
        }
    un = paired["un"].to_numpy(dtype=float)
    he = paired["he"].to_numpy(dtype=float)
    notion = paired["notional"].to_numpy(dtype=float)
    imp = np.abs(he) < np.abs(un)
    rmse_un = float(np.sqrt(np.mean(un ** 2)))
    rmse_he = float(np.sqrt(np.mean(he ** 2)))
    mae_un = float(np.mean(np.abs(un)))
    mae_he = float(np.mean(np.abs(he)))
    var_un = float(np.var(un, ddof=1)) if len(un) > 1 else np.nan
    var_he = float(np.var(he, ddof=1)) if len(he) > 1 else np.nan
    q05, es05 = _tail_stats(paired["he"])
    return {
        "n_intervals": int(len(paired)),
        "mean_unhedged_coin": float(np.mean(un)),
        "mean_hedged_coin": float(np.mean(he)),
        "std_unhedged_coin": float(np.std(un, ddof=1)) if len(un) > 1 else 0.0,
        "std_hedged_coin": float(np.std(he, ddof=1)) if len(he) > 1 else 0.0,
        "var_unhedged_coin": var_un,
        "var_hedged_coin": var_he,
        "rmse_unhedged_coin": rmse_un,
        "rmse_hedged_coin": rmse_he,
        "rmse_reduction": 1.0 - (rmse_he / rmse_un) if rmse_un > 0 else np.nan,
        "mae_unhedged_coin": mae_un,
        "mae_hedged_coin": mae_he,
        "mae_reduction": 1.0 - (mae_he / mae_un) if mae_un > 0 else np.nan,
        "variance_reduction": 1.0 - (var_he / var_un) if np.isfinite(var_un) and var_un > 0 and np.isfinite(var_he) else np.nan,
        "hit_rate_abs_improvement": float(np.mean(imp)),
        "q05_hedged_coin": q05,
        "es05_hedged_coin": es05,
        "mean_abs_hedge_notional_usd": float(np.nanmean(np.abs(notion))),
    }


def summarize_panel(panel: pd.DataFrame, group_cols: list[str]) -> pd.DataFrame:
    rows = []
    eligible = panel[panel["is_summary_eligible"].fillna(False)].copy()
    for model in MODELS:
        for hedge_type in HEDGE_TYPES:
            he_col = f"net_pnl_coin_{model}_{hedge_type}"
            notion_col = f"hedge_notional_usd_{model}_{hedge_type}"
            for keys, g in eligible.groupby(group_cols, dropna=False):
                if not isinstance(keys, tuple):
                    keys = (keys,)
                row = {col: val for col, val in zip(group_cols, keys)}
                row.update({"model": model, "hedge_type": hedge_type})
                row.update(_summarize_errors(g["option_pnl_coin_short"], g[he_col], g[notion_col]))
                rows.append(row)
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    sort_cols = [c for c in group_cols if c in out.columns] + ["model", "hedge_type"]
    return out.sort_values(sort_cols).reset_index(drop=True)


def make_summary_tables_from_panel(panel: pd.DataFrame) -> dict[str, pd.DataFrame]:
    panel_b = add_analysis_buckets(panel)
    return {
        "hedge_summary_overall": summarize_panel(panel_b, ["currency", "split"]),
        "hedge_summary_by_timestamp": summarize_panel(panel_b, ["snapshot_ts", "currency", "split"]),
        "hedge_summary_by_option": summarize_panel(panel_b, ["currency", "split", "instrument_name"]),
        "hedge_summary_by_maturity_bucket": summarize_panel(panel_b, ["currency", "split", "maturity_bucket"]),
        "hedge_summary_by_moneyness_bucket": summarize_panel(panel_b, ["currency", "split", "moneyness_bucket"]),
    }


def prepare_output_option_sheets(options: pd.DataFrame, workbook: dict[str, pd.DataFrame]) -> tuple[pd.DataFrame, pd.DataFrame]:
    base_train_cols = list(workbook["train_data"].columns)
    base_test_cols = list(workbook["test_data"].columns)
    enriched = options.copy()
    train = enriched[enriched["split"] == "train"].copy()
    test = enriched[enriched["split"] == "test"].copy()
    train_out = train[base_train_cols + DELTA_OUTPUT_COLS].copy()
    test_out = test[base_test_cols + DELTA_OUTPUT_COLS].copy()
    return train_out, test_out


def make_notes(cfg: HedgingConfig, panel: pd.DataFrame, perp: pd.DataFrame, funding: pd.DataFrame, used_perp_file: str, used_funding_file: str | None) -> pd.DataFrame:
    rows = [
        ("calibration_xlsx", str(cfg.calibration_xlsx)),
        ("data_dir", str(cfg.data_dir)),
        ("output_xlsx", str(cfg.output_xlsx)),
        ("hedge_price_file", used_perp_file),
        ("funding_file", used_funding_file),
        ("dynamic_b", cfg.dynamic_b),
        ("verify_prices", cfg.verify_prices),
        ("rebalance_every_n_snapshots", cfg.rebalance_every_n_snapshots),
        ("fft_N", cfg.fft_params.N),
        ("fft_eta", cfg.fft_params.eta),
        ("fft_alpha", cfg.fft_params.alpha),
        ("delta_u_max", cfg.delta_u_max),
        ("delta_quad_n", cfg.delta_quad_n),
        ("max_snapshot_groups", cfg.max_snapshot_groups),
        ("currencies", None if cfg.currencies is None else ",".join(cfg.currencies)),
        ("delta_definition", "regular_delta = dV_usd/dF"),
        ("net_delta_definition", "net_delta = F * dV_coin/dF = regular_delta - V_coin"),
        ("inverse_perp_pnl_definition", "hedge_pnl_coin = N_usd * (1/H_t - 1/H_t1)"),
        ("delta_hedge_ratio_definition", "N_usd = H_t * regular_delta"),
        ("net_delta_hedge_ratio_definition", "N_usd = H_t^2 / F_t * net_delta"),
        ("funding_note", "Funding is summarized but excluded from hedged net PnL metrics."),
        ("exclude_expiry_crossing", cfg.exclude_expiry_crossing),
        ("interval_panel_rows", int(len(panel))),
        ("n_unique_snapshots", int(panel["snapshot_ts"].nunique()) if not panel.empty else 0),
        ("n_unique_instruments", int(panel["instrument_name"].nunique()) if not panel.empty else 0),
        ("perp_rows", int(len(perp))),
        ("funding_rows", int(len(funding))),
    ]
    return pd.DataFrame(rows, columns=["item", "value"])


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells[:1000])
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)


def _write_df_to_sheet_chunked(wb: Workbook, sheet_name: str, df: pd.DataFrame) -> None:
    df = _stringify_datetimes(df)
    total_rows = len(df)
    if total_rows <= EXCEL_MAX_DATA_ROWS:
        ws = wb.create_sheet(sheet_name[:31])
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
        _style_sheet(ws)
        return
    n_parts = int(math.ceil(total_rows / EXCEL_MAX_DATA_ROWS))
    for part in range(n_parts):
        start = part * EXCEL_MAX_DATA_ROWS
        end = min((part + 1) * EXCEL_MAX_DATA_ROWS, total_rows)
        ws = wb.create_sheet(f"{sheet_name}_{part + 1}"[:31])
        for row in dataframe_to_rows(df.iloc[start:end], index=False, header=True):
            ws.append(row)
        _style_sheet(ws)


def write_hedge_workbook(output_path: Path, workbook: dict[str, pd.DataFrame], train_data: pd.DataFrame, test_data: pd.DataFrame, price_verification: pd.DataFrame, interval_panel: pd.DataFrame, summary_tables: dict[str, pd.DataFrame], notes: pd.DataFrame) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ordered = [
        ("black_params", workbook["black_params"].copy()),
        ("heston_params", workbook["heston_params"].copy()),
        ("svcj_params", workbook["svcj_params"].copy()),
        ("train_data", train_data),
        ("test_data", test_data),
        ("hedge_interval_panel", interval_panel),
        ("hedge_summary_overall", summary_tables["hedge_summary_overall"]),
        ("hedge_summary_by_timestamp", summary_tables["hedge_summary_by_timestamp"]),
        ("hedge_summary_by_option", summary_tables["hedge_summary_by_option"]),
        ("hedge_summary_by_maturity_bucket", summary_tables["hedge_summary_by_maturity_bucket"]),
        ("hedge_summary_by_moneyness_bucket", summary_tables["hedge_summary_by_moneyness_bucket"]),
        ("price_verification", price_verification),
        ("hedge_notes", notes),
    ]
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    for sheet_name, df in ordered:
        _write_df_to_sheet_chunked(wb, sheet_name, df)
    wb.save(output_path)


def run_hedging_analysis(cfg: HedgingConfig) -> dict[str, pd.DataFrame]:
    outer = _make_progress(8, "run_hedging_analysis")
    try:
        outer.set_description("load workbook")
        workbook = load_calibration_workbook(cfg.calibration_xlsx)
        outer.update(1)

        outer.set_description("combine options")
        options = filter_options_for_run(combine_option_sheets(workbook), cfg)
        param_lookup = build_param_lookup(workbook)
        outer.update(1)

        outer.set_description("reprice and deltas")
        n_blocks = int(options.groupby(["snapshot_ts", "currency"]).ngroups) * len(MODELS)
        inner = _make_progress(n_blocks, "model blocks")
        try:
            options, price_verification = enrich_options_with_model_outputs(
                options,
                param_lookup,
                fft_params=cfg.fft_params,
                dynamic_b=cfg.dynamic_b,
                verify_prices=cfg.verify_prices,
                delta_u_max=cfg.delta_u_max,
                delta_quad_n=cfg.delta_quad_n,
                progress=inner,
            )
        finally:
            inner.close()
        outer.update(1)

        outer.set_description("attach evaluation snapshot")
        snapshot_index = list_snapshot_files(cfg.data_dir)
        options = attach_evaluation_snapshot_data(options, snapshot_index, rebalance_every_n_snapshots=cfg.rebalance_every_n_snapshots)
        outer.update(1)

        outer.set_description("attach perp and funding")
        perp, used_perp_file = load_perp_history(cfg.data_dir, cfg.hedge_price_file)
        funding, used_funding_file = load_funding_history(cfg.data_dir, cfg.funding_file)
        options = attach_perp_and_funding(options, perp, funding)
        outer.update(1)

        outer.set_description("build interval panel")
        interval_panel = build_hedge_interval_panel(options, exclude_expiry_crossing=cfg.exclude_expiry_crossing, rebalance_every_n_snapshots=cfg.rebalance_every_n_snapshots)
        interval_panel = add_analysis_buckets(interval_panel)
        outer.update(1)

        outer.set_description("summaries")
        summary_tables = make_summary_tables_from_panel(interval_panel)
        train_data, test_data = prepare_output_option_sheets(options, workbook)
        notes = make_notes(cfg, interval_panel, perp, funding, used_perp_file, used_funding_file)
        outer.update(1)

        outer.set_description("write workbook")
        write_hedge_workbook(cfg.output_xlsx, workbook, train_data, test_data, price_verification, interval_panel, summary_tables, notes)
        outer.update(1)

        return {
            "train_data": train_data,
            "test_data": test_data,
            "price_verification": price_verification,
            "hedge_interval_panel": interval_panel,
            **summary_tables,
            "hedge_notes": notes,
        }
    finally:
        outer.close()